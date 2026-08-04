# -*- coding: utf-8 -*-
"""정규화 레코드 리스트 -> 스타일이 적용된 엑셀 파일 생성."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# 출력 컬럼 순서 (사용자 요청: 공종·적용기준일·공사규모·공사기간·비목·요율 + 보조컬럼)
COLUMNS = [
    ("공종", 8),
    ("적용기준일", 12),
    ("비목", 24),
    ("산정기준", 20),
    ("세부구분", 34),
    ("공사규모", 16),
    ("공사기간", 18),
    ("요율(%)", 10),
    ("기초액(천원)", 12),
    ("비고", 40),
]

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
THIN = Side(style="thin", color="D0D0D0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 비목 표준 정렬 순서
BIMOK_ORDER = [
    "간접노무비", "기타경비", "일반관리비", "이윤",
    "산업안전보건관리비", "건설기계대여대금지급보증액",
    "건강보험료", "노인장기요양보험료", "연금보험료", "산재보험료", "고용보험료",
    "환경보전비", "퇴직공제부금비",
    "건설하도급대금지급보증서발급수수료", "공사이행보증수수료",
    "법정부담금(석면분담금)", "법정부담금(임금채권부담금)",
]
_ORDER = {n: i for i, n in enumerate(BIMOK_ORDER)}


def _sort_key(r):
    return (
        r.get("공종", ""),
        r.get("적용기준일", ""),
        _ORDER.get(r.get("비목", ""), 99),
        str(r.get("세부구분", "")),
        str(r.get("공사규모", "")),
        str(r.get("공사기간", "")),
    )


def build_workbook(records):
    """records -> openpyxl Workbook (정규화_전체 시트 + 요약 시트)."""
    records = sorted(records, key=_sort_key)
    wb = openpyxl.Workbook()

    # --- 시트1: 정규화 전체 ---
    ws = wb.active
    ws.title = "정규화_전체"
    headers = [c[0] for c in COLUMNS]
    ws.append(headers)
    for i, (name, width) in enumerate(COLUMNS, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER

    num_cols = {headers.index("요율(%)") + 1, headers.index("기초액(천원)") + 1}
    for r in records:
        row = [r.get(h) for h in headers]
        ws.append(row)
        rr = ws.max_row
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=rr, column=ci)
            cell.border = BORDER
            if ci in num_cols:
                cell.alignment = CENTER
                if isinstance(cell.value, (int, float)):
                    cell.number_format = "0.####"
            elif headers[ci - 1] in ("공종", "적용기준일", "공사규모", "공사기간"):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)), ws.max_row)

    # --- 시트2: 요약(공종·적용기준일·비목별 건수) ---
    ws2 = wb.create_sheet("요약")
    from collections import Counter
    cnt = Counter((r.get("공종"), r.get("적용기준일"), r.get("비목")) for r in records)
    ws2.append(["공종", "적용기준일", "비목", "건수"])
    for i in range(1, 5):
        c = ws2.cell(row=1, column=i)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
    for (g, d, b), n in sorted(cnt.items(),
                               key=lambda x: (x[0][0] or "", x[0][1] or "",
                                              _ORDER.get(x[0][2], 99))):
        ws2.append([g, d, b, n])
    for i, w in enumerate([8, 12, 26, 8], start=1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    return wb


def export(records, path):
    wb = build_workbook(records)
    wb.save(path)
    return path
