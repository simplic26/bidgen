# -*- coding: utf-8 -*-
"""
조달청 시설공사 원가계산 간접공사비(제비율) 적용기준 엑셀 파서.

토목/건축 두 가지 템플릿(공종)을 위치 기반으로 파싱하여
공종, 적용기준일, 비목, 산정기준, 세부구분, 공사규모, 공사기간, 요율(%), 기초액, 비고
형태의 정규화 레코드 리스트로 변환한다.

레이아웃은 날짜가 달라도 동일하며, 토목/건축은 '컬럼 위치'만 다르고
'행 위치'는 대부분 공유한다. 따라서 행 정의는 공유하고 컬럼맵만 공종별로 둔다.
"""
import re
import os
import zipfile
import datetime
from xml.etree import ElementTree as ET

import openpyxl
from openpyxl.utils import (column_index_from_string, get_column_letter,
                            range_boundaries)

_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
_R = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"


# ---------------------------------------------------------------------------
# 빠른 로더: read_only 로 값 그리드 + zip XML 에서 병합범위 직접 파싱
#   (openpyxl 일반 로드는 파일당 3초 → read_only 는 0.2초, 20배 빠름)
# ---------------------------------------------------------------------------
def _first_sheet_xml_path(z):
    """workbook.xml + rels 로 '첫 번째 시트'의 XML 경로를 찾는다."""
    try:
        wb = ET.fromstring(z.read("xl/workbook.xml"))
        first = wb.find("m:sheets/m:sheet", _NS)
        rid = first.get(_R)
        rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
        for rel in rels:
            if rel.get("Id") == rid:
                tgt = rel.get("Target")
                if tgt.startswith("/"):
                    return tgt.lstrip("/")
                return "xl/" + tgt.replace("../", "").lstrip("./")
    except Exception:  # noqa
        pass
    return "xl/worksheets/sheet1.xml"


def _read_merges(path):
    """xlsx 첫 시트의 병합범위 목록 [(min_row,min_col,max_row,max_col)]."""
    out = []
    with zipfile.ZipFile(path) as z:
        sp = _first_sheet_xml_path(z)
        try:
            data = z.read(sp)
        except KeyError:
            return out
        root = ET.fromstring(data)
        mc = root.find("m:mergeCells", _NS)
        if mc is None:
            return out
        for m in mc.findall("m:mergeCell", _NS):
            ref = m.get("ref")
            if ref:
                min_c, min_r, max_c, max_r = range_boundaries(ref)
                out.append((min_r, min_c, max_r, max_c))
    return out


def load_fast(path):
    """(grid, merges) 반환. grid[(row,col)] = 값(비어있으면 없음)."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb.worksheets[0]
    grid = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                grid[(cell.row, cell.column)] = cell.value
    wb.close()
    merges = _read_merges(path)
    return grid, merges


# ---------------------------------------------------------------------------
# 셀 리더 (병합셀 해석)
# ---------------------------------------------------------------------------
class SheetReader:
    """값 그리드 + 병합범위로 셀 값을 돌려주는 리더 (병합 자동 해석)."""

    def __init__(self, grid, merges):
        self.grid = grid
        # (row, col) -> (anchor_row, anchor_col)  병합범위 대표셀 매핑
        self._merge = {}
        for (min_r, min_c, max_r, max_c) in merges:
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    self._merge[(r, c)] = (min_r, min_c)

    def raw(self, row, col):
        """병합 대표셀 값(원본)."""
        a = self._merge.get((row, col))
        if a:
            row, col = a
        return self.grid.get((row, col))

    def is_anchor(self, col_letter, row):
        """해당 셀이 (병합되지 않았거나) 병합범위의 대표(좌상단)행인지."""
        col = column_index_from_string(col_letter)
        a = self._merge.get((row, col))
        return (a is None) or (a[0] == row)

    def get(self, ref):
        """'AC15' 같은 좌표 문자열로 값 조회 (병합 해석)."""
        m = re.match(r"([A-Z]+)(\d+)", ref)
        col = column_index_from_string(m.group(1))
        row = int(m.group(2))
        return self.raw(row, col)

    def cell(self, col_letter, row):
        return self.raw(row, column_index_from_string(col_letter))

    def text(self, col_letter, row):
        v = self.raw(row, column_index_from_string(col_letter))
        if v is None:
            return ""
        return re.sub(r"\s+", " ", str(v)).strip()


# ---------------------------------------------------------------------------
# 값 변환 유틸
# ---------------------------------------------------------------------------
def to_number(v):
    """숫자/숫자문자열을 float로. '(1.8)', '1.30', '1,234' 등 허용. 실패시 None."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    neg = False
    if s.startswith("(") and s.endswith(")"):
        s = s[1:-1].strip()  # 괄호값(예시/참고치)은 값으로 취급
    s = s.replace(",", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None


def first_number_in(text):
    """문자열 안의 첫 실수를 추출 (예: '(직노) x 3.595' -> 3.595)."""
    if text is None:
        return None
    m = re.search(r"(\d+(?:\.\d+)?)", str(text))
    return float(m.group(1)) if m else None


def clean(text):
    if text is None:
        return ""
    return re.sub(r"\s+", " ", str(text)).strip()


# ---------------------------------------------------------------------------
# 파일명 파싱: 공종 + 적용기준일
# ---------------------------------------------------------------------------
GONGJONG_MAP = [
    ("토목", "토목"),
    ("건축", "건축"),
]


def parse_filename(fname):
    """파일명에서 (공종, 적용기준일 datetime.date) 추출.
    예: '토목공사 간접공사비 적용기준(260413).xlsx' -> ('토목', 2026-04-13)
    """
    base = os.path.basename(fname)
    gongjong = None
    for key, label in GONGJONG_MAP:
        if key in base:
            gongjong = label
            break
    m = re.search(r"\((\d{6})\)", base)
    apply_date = None
    if m:
        s = m.group(1)
        yy, mm, dd = int(s[0:2]), int(s[2:4]), int(s[4:6])
        year = 2000 + yy
        try:
            apply_date = datetime.date(year, mm, dd)
        except ValueError:
            apply_date = None
    return gongjong, apply_date


# ---------------------------------------------------------------------------
# 행 정의 (토목/건축 공유)
# ---------------------------------------------------------------------------
# 간접노무비/기타경비 매트릭스: 공사규모 5단계 x 공사기간 4단계
MATRIX_ROWS = [15, 17, 19, 21, 23, 25, 27, 29, 31, 33, 35, 37, 39, 41, 43, 45, 47, 49, 51, 53]
# 일반관리비/이윤: 공사규모(추정가격) 앵커 행
GENERAL_ROWS = [15, 21, 27, 31, 35, 39, 43]


# ---------------------------------------------------------------------------
# 공종별 컬럼맵
# ---------------------------------------------------------------------------
# 각 값은 '엑셀 컬럼 문자'.
TEMPLATES = {
    "토목": {
        "sub_labor": [("토목", "AC"), ("조경", "AF"), ("산업설비(토목)", "AI")],
        "sub_etc":   [("토목", "AL"), ("조경", "AO"), ("산업설비(토목)", "AR")],
        "gen_scale_col": "AU",
        "gen_cols": [("토목·조경·산업설비(토목)", "BD"), ("전문·전기·통신·소방·기타", "BL")],
        "profit_col": "BS",
        # 건설기계대여대금 지급보증액
        "mach_general": ("BZ", "CI"),   # (구분, 요율)
        "mach_pro":     ("CN", "DE"),
        # 산업안전보건관리비
        "safety": {"target": "DK", "kind": "EB", "rate": "ES", "base": "FA", "sub2": "DR"},
        # 사회보험 (고정식) : 값이 들어있는 셀
        "health": "AU51", "ltc": "BD51", "pension": "BR51", "accident": "BZ51",
        # 고용보험료 : 등급/요율
        "employ": ("B", "AH"),
        # 환경보전비 : (그룹, 구분, 요율) — 토목은 그룹(AU)+세부(AZ)
        "env": {"group": "AU", "kind": "AZ", "rate": "BU"},
        # 건설하도급대금 지급보증서 발급수수료
        "subcon": ("BZ", "DC"),
        # 공사이행보증수수료 : 규모/수수료(수식문자)
        "perf": ("B", "P"),
        # 법정부담금
        "asbestos": "BZ99", "wageclaim": "CS99",
        # 퇴직공제부금비
        "retire": "AU110",
    },
    "건축": {
        "sub_labor": [("건축", "AC"), ("산업설비(건축)", "AH")],
        "sub_etc":   [("건축", "AL"), ("산업설비(건축)", "AQ")],
        "gen_scale_col": "AU",
        "gen_cols": [("건축·산업설비(건축)", "BE"), ("전문·전기·통신·소방·기타", "BK")],
        "profit_col": "BR",
        "mach_general": ("BY", "CH"),
        "mach_pro":     ("CM", "DD"),
        "safety": {"target": "DJ", "kind": "EA", "rate": "ER", "base": "EZ", "sub2": "DQ"},
        "health": "AU51", "ltc": "BC51", "pension": "BQ51", "accident": "BY51",
        "employ": ("B", "AH"),
        "env": {"group": None, "kind": "AU", "rate": "BT"},   # 건축은 구분(AU) 단일
        "subcon": ("BY", "DB"),
        "perf": ("B", "P"),
        "asbestos": "BY99", "wageclaim": "CR99",
        "retire": "AU110",
    },
}


# ---------------------------------------------------------------------------
# 메인 파서
# ---------------------------------------------------------------------------
def _rec(gongjong, apply_date, 비목, 산정기준, 세부구분, 규모, 기간, 요율, 기초액=None, 비고=""):
    return {
        "공종": gongjong,
        "적용기준일": apply_date.isoformat() if apply_date else "",
        "비목": 비목,
        "산정기준": 산정기준,
        "세부구분": 세부구분,
        "공사규모": 규모,
        "공사기간": 기간,
        "요율(%)": 요율,
        "기초액(천원)": 기초액,
        "비고": 비고,
    }


def parse_workbook(path):
    """엑셀 1개 파일 -> 정규화 레코드 리스트."""
    gongjong, apply_date = parse_filename(path)
    if gongjong not in TEMPLATES:
        raise ValueError("지원하지 않는 공종(파일명에 '토목' 또는 '건축' 필요): %s" % path)
    tpl = TEMPLATES[gongjong]

    grid, merges = load_fast(path)
    r = SheetReader(grid, merges)
    recs = []

    def R(*a, **k):
        recs.append(_rec(gongjong, apply_date, *a, **k))

    # 1) 간접노무비  (직노) x 율 : 규모 x 기간 x 세부공종
    for label, col in tpl["sub_labor"]:
        for row in MATRIX_ROWS:
            val = to_number(r.cell(col, row))
            if val is None:
                continue
            R("간접노무비", "(직노) x 율", label,
              r.text("B", row), r.text("M", row), val)

    # 2) 기타경비  (재+노) x 율
    for label, col in tpl["sub_etc"]:
        for row in MATRIX_ROWS:
            val = to_number(r.cell(col, row))
            if val is None:
                continue
            R("기타경비", "(재+노) x 율", label,
              r.text("B", row), r.text("M", row), val)

    # 3) 일반관리비  (재+노+경) x 율 : 규모별
    for label, col in tpl["gen_cols"]:
        for row in GENERAL_ROWS:
            val = to_number(r.cell(col, row))
            if val is None:
                continue
            R("일반관리비", "(재+노+경) x 율", label,
              r.text(tpl["gen_scale_col"], row), "", val)

    # 4) 이윤  (노+경+일) x 율 : 규모별
    for row in GENERAL_ROWS:
        val = to_number(r.cell(tpl["profit_col"], row))
        if val is None:
            continue
        R("이윤", "(노+경+일) x 율", "",
          r.text(tpl["gen_scale_col"], row), "", val)

    # 5) 건설기계대여대금 지급보증액  (직접공사비) x 율
    for gubun, (kcol, vcol) in [("종합건설업", tpl["mach_general"]),
                                 ("전문건설업", tpl["mach_pro"])]:
        for row in range(17, 41):
            if not r.is_anchor(kcol, row):
                continue
            kind = r.text(kcol, row)
            val = to_number(r.cell(vcol, row))
            if val is None or not kind or kind.startswith("[") or "구분" in kind:
                continue
            R("건설기계대여대금지급보증액", "(직접공사비) x 율", "%s / %s" % (gubun, kind),
              "", "", val)

    # 6) 산업안전보건관리비  (재+직노) x 요율 + 기초액
    sf = tpl["safety"]
    for row in range(17, 52):
        if not r.is_anchor(sf["kind"], row):
            continue
        kind = r.text(sf["kind"], row)
        val = to_number(r.cell(sf["rate"], row))
        if val is None or not kind or "구분" in kind:
            continue
        target = r.text(sf["target"], row)
        base = to_number(r.cell(sf["base"], row))
        note = "대상액: %s" % target if target else ""
        sub2col = sf.get("sub2")
        if sub2col:
            sub2 = r.text(sub2col, row)
            if sub2 and sub2 != target:
                note = (note + " / " + sub2).strip(" /")
        R("산업안전보건관리비", "(재+직노) x 요율 + 기초액", kind,
          "", "", val, 기초액=base, 비고=note)

    # 7) 사회보험료 (고정 산정식)
    for 비목, key, basis in [
        ("건강보험료", "health", "(직노) x 율"),
        ("노인장기요양보험료", "ltc", "(건강보험료) x 율"),
        ("연금보험료", "pension", "(직노) x 율"),
        ("산재보험료", "accident", "(노) x 율"),
    ]:
        raw = clean(r.get(tpl[key]))
        val = first_number_in(raw)
        if val is not None:
            R(비목, basis, "", "", "", val, 비고=raw)

    # 8) 고용보험료  (노) x 율 : 공사배정규모(등급)별
    ecol, evcol = tpl["employ"]
    for row in range(66, 83):
        if not r.is_anchor(ecol, row) or not r.is_anchor(evcol, row):
            continue
        grade = r.text(ecol, row)
        val = to_number(r.cell(evcol, row))
        if val is None or not grade or "등급" not in grade:
            continue
        R("고용보험료", "(노) x 율", clean(grade), "", "", val)

    # 9) 환경보전비  (직접공사비) x 율 : 공사종류별
    env = tpl["env"]
    for row in range(66, 96):
        if not r.is_anchor(env["kind"], row):
            continue
        kind = r.text(env["kind"], row)
        val = to_number(r.cell(env["rate"], row))
        if val is None or not kind or kind.startswith("[") or "구분" in kind or "종류" in kind:
            continue
        gubun = kind
        if env["group"]:
            grp = r.text(env["group"], row)
            if grp and grp != kind and not grp.startswith("*") and not grp.startswith("※"):
                gubun = "%s / %s" % (grp, kind)
        R("환경보전비", "(직접공사비) x 율", gubun, "", "", val)

    # 10) 건설하도급대금 지급보증서 발급수수료  (직접공사비) x 율 : 공사규모별
    scol, svcol = tpl["subcon"]
    for row in range(66, 91):
        if not r.is_anchor(scol, row):
            continue
        scale = r.text(scol, row)
        val = to_number(r.cell(svcol, row))
        if val is None or not scale or scale.startswith("[") or "규모" in scale:
            continue
        R("건설하도급대금지급보증서발급수수료", "(직접공사비) x 율", "", clean(scale), "", val)

    # 11) 공사이행보증수수료 : 규모별 수수료(수식) — 요율이 수식이므로 비고에 보관
    pcol, pvcol = tpl["perf"]
    for row in range(96, 112):
        if not r.is_anchor(pcol, row):
            continue
        scale = r.text(pcol, row)
        formula = r.text(pvcol, row)
        if not scale or not formula or "규모" in scale or "수수료" in formula:
            continue
        if not re.search(r"\d", scale):
            continue
        R("공사이행보증수수료", "공사규모별 수수료율", "", clean(scale), "", None, 비고=clean(formula))

    # 12) 법정부담금
    for 비목, key in [("법정부담금(석면분담금)", "asbestos"),
                       ("법정부담금(임금채권부담금)", "wageclaim")]:
        raw = clean(r.get(tpl[key]))
        val = first_number_in(raw)
        if val is not None:
            basis = raw.split("x")[0].strip() + " x 율" if "x" in raw else raw
            R(비목, basis, "", "", "", val, 비고=raw)

    # 13) 퇴직공제부금비
    raw = clean(r.get(tpl["retire"]))
    val = first_number_in(raw)
    if val is not None:
        R("퇴직공제부금비", "(직노) x 율", "", "", "", val, 비고=raw)

    return recs


def parse_files(paths):
    """여러 파일 -> 통합 레코드 리스트 (+ 오류 목록)."""
    all_recs, errors = [], []
    for p in paths:
        try:
            all_recs.extend(parse_workbook(p))
        except Exception as e:  # noqa
            errors.append((os.path.basename(p), str(e)))
    return all_recs, errors


if __name__ == "__main__":
    import sys
    import json
    recs, errs = parse_files(sys.argv[1:])
    print("records:", len(recs), "errors:", errs)
    print(json.dumps(recs[:5], ensure_ascii=False, indent=1))
